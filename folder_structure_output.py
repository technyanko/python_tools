import os
from openpyxl import Workbook
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

# エクセルファイルを保存する関数
def save_excel(folder_path, progress_bar):
    # フォルダパスが空の場合
    if not folder_path:
        return

    # 出力ワークブックとシートを作成
    wb = Workbook()
    ws = wb.active

    # ルートディレクトリ設定
    root_dir = folder_path

    # シート名にルートディレクトリ名を設定
    ws.title = os.path.basename(root_dir)

    # フォント設定（赤色、太字）
    red_bold_font = Font(color="FF0000", bold=True)
    black_bold_font = Font(bold=True)

    # エクセルのA1セルにルートディレクトリ名を出力
    cell_a1 = ws['A1']
    cell_a1.value = os.path.basename(root_dir) + 'のフォルダ構成一覧'
    cell_a1.font = black_bold_font  # フォントを黒色、太字に設定

    # ファイルやフォルダの総数を取得
    total_entries = count_entries(root_dir)

    # 進捗バーの設定
    progress_bar["maximum"] = total_entries
    progress_bar["value"] = 0
    progress_bar.update()

    # 再帰的にディレクトリを探索
    def explore_dir(current_dir, indent=0):
        entries_processed = 0
        for entry in os.listdir(current_dir):
            entry_path = os.path.join(current_dir, entry)
            if os.path.isdir(entry_path):
                # フォルダの場合はインデントを増やして再帰的に探索
                cell = ws.cell(row=ws.max_row + 1, column=indent + 1)
                cell.value = entry
                cell.font = red_bold_font  # フォントを赤色、太字に設定
                entries_processed += 1
                progress_bar["value"] += 1
                progress_bar.update()
                explore_dir(entry_path, indent + 1)
            else:
                # ファイルの場合は隣の列に出力
                ws.cell(row=ws.max_row + 1, column=indent + 1).value = entry
                entries_processed += 1
                progress_bar["value"] += 1
                progress_bar.update()

    # ルートディレクトリを探索開始
    explore_dir(root_dir)

    # エクセルファイルを保存
    excel_filename = os.path.basename(root_dir) + '_フォルダ構成.xlsx'
    wb.save(excel_filename)

    # メッセージをポップアップで表示
    messagebox.showinfo("完了", f"フォルダ構成をエクセルファイル「{excel_filename}」に出力しました。")

# フォルダ内のファイルやフォルダの数を再帰的に取得
def count_entries(folder_path):
    count = 0
    for root, dirs, files in os.walk(folder_path):
        count += len(dirs) + len(files)
    return count

# フォルダ選択ダイアログを開いてフォルダパスを取得
def select_folder():
    folder_path = filedialog.askdirectory(title='出力対象フォルダを選択してください')
    if folder_path:
        save_button.config(state=tk.NORMAL)
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, folder_path)

# フォルダパスの入力欄を作成
def create_folder_input():
    folder_frame = ttk.Frame(root)
    folder_frame.pack(fill="x", padx=10, pady=5)

    folder_entry_label = ttk.Label(folder_frame, text="出力対象フォルダパス:")
    folder_entry_label.pack(side="left")

    folder_entry = ttk.Entry(folder_frame)
    folder_entry.pack(side="left", expand=True, fill="x")

    folder_button = ttk.Button(folder_frame, text="参照", command=select_folder)
    folder_button.pack(side="left")

    return folder_entry

# ポップアップウィンドウを作成
root = tk.Tk()
root.title('フォルダ構成出力')

# フォルダパスの入力欄を作成
folder_entry = create_folder_input()

# 進捗バーを作成
progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=5)

# 実行ボタンを作成
save_button = ttk.Button(root, text="実行", command=lambda: save_excel(folder_entry.get(), progress_bar), state=tk.DISABLED)
save_button.pack(pady=5)

# メインループを実行
root.mainloop()
